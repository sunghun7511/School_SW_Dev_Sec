# -*- coding: utf-8 -*-

from openpyxl import load_workbook

# load_workbook 함수를 이용하여 엑셀 클래스 변수 생성
wb = load_workbook("test01.xlsx")
# 활성화 된 시트를 sheet 변수로 설정
sheet = wb.active

print(sheet["A1"].value)
print(sheet["A2"].value)
print(sheet["B1"].value)
print(sheet["C1"].value)
print(sheet["D3"].value)
