from MailFunction import Email
from openpyxl import load_workbook

wb = load_workbook("mails.xlsx")
sheet = wb.active

e = Email()

for row in sheet.iter_rows():
    name = row[0].value
    email = row[1].value

    e.send_mail(name, email, subject="희희흐히ㅡ히희희흐힣ㅎ", content="재웅쌤 사랑해요 흐히희ㅡ히ㅣㅎㅎ")